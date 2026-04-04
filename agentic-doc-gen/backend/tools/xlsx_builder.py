"""
XLSX 문서 빌드 엔진

DocumentOutput (JSON 스키마) → XLSX 파일 변환.
document_schema.py의 DocumentOutput을 입력받아 openpyxl로 Excel 파일을 생성한다.

각 섹션의 table, heading, paragraph, list 요소를 단일 시트에 순차 배치한다.
"""

import logging
import tempfile
import traceback

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..document_schema import (
    DocumentOutput,
    ListContent,
    TableContent,
    TextContent,
)

logger = logging.getLogger(__name__)


# ── 스타일 상수 ──

# 문서 제목
_TITLE_FONT = Font(bold=True, size=14)
_TITLE_ALIGN = Alignment(horizontal="center", vertical="center")

# 섹션/heading 스타일
_HEADING_FONT = Font(bold=True, size=12)

# 테이블 헤더
_TABLE_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_TABLE_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TABLE_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

# 테이블 데이터 교차 행
_DATA_FILL_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_DATA_FILL_ODD = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
_DATA_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# 보더
_THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

# paragraph 스타일
_PARAGRAPH_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# list 스타일
_LIST_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 기본 열 너비 기준값 (비율 100 기준 → 실제 Excel 너비)
_BASE_WIDTH_FACTOR = 0.5
_MIN_COL_WIDTH = 8
_MAX_COL_WIDTH = 60


# ── 정렬 매핑 ──

_ALIGN_MAP = {
    "left": _DATA_ALIGN_LEFT,
    "center": _DATA_ALIGN_CENTER,
    "right": _DATA_ALIGN_RIGHT,
}


def _calc_column_widths(columns) -> list[float]:
    """TableColumn 리스트에서 Excel 열 너비를 계산한다.

    columns의 width가 비율(합계 100 기준)이면 비율 기반으로,
    없으면 열 이름 길이 기반으로 자동 계산한다.
    """
    widths = []
    for col in columns:
        if col.width and col.width > 0:
            # 비율 기반: width=30 → 약 15 정도 Excel 너비
            w = max(col.width * _BASE_WIDTH_FACTOR, _MIN_COL_WIDTH)
            widths.append(min(w, _MAX_COL_WIDTH))
        else:
            # 열 이름 길이 기반
            name_len = len(col.name) * 2 + 4  # 한글 고려 (2바이트)
            widths.append(max(name_len, _MIN_COL_WIDTH))
    return widths


def _write_title(ws, title: str, num_cols: int, current_row: int) -> int:
    """문서 제목을 첫 행에 병합 셀로 작성한다. 다음 행 번호를 반환."""
    if num_cols > 1:
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=num_cols,
        )
    cell = ws.cell(row=current_row, column=1, value=title)
    cell.font = _TITLE_FONT
    cell.alignment = _TITLE_ALIGN
    return current_row + 1


def _write_heading(ws, content: TextContent, current_row: int, max_cols: int) -> int:
    """heading 요소를 볼드 텍스트 행으로 작성한다."""
    cell = ws.cell(row=current_row, column=1, value=content.text)
    font_size = content.font_size or 12
    cell.font = Font(bold=True, size=font_size)
    if content.alignment:
        cell.alignment = Alignment(
            horizontal=content.alignment, vertical="center",
        )
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center")
    # heading이 여러 열에 걸쳐야 하면 병합
    if max_cols > 1:
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=max_cols,
        )
    return current_row + 1


def _write_paragraph(ws, content: TextContent, current_row: int, max_cols: int) -> int:
    """paragraph 요소를 일반 텍스트 행으로 작성한다 (워드랩 적용)."""
    cell = ws.cell(row=current_row, column=1, value=content.text)
    font_size = content.font_size or 10
    cell.font = Font(bold=content.bold, size=font_size)
    cell.alignment = _PARAGRAPH_ALIGN
    if max_cols > 1:
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=max_cols,
        )
    return current_row + 1


def _write_table(ws, content: TableContent, current_row: int) -> int:
    """테이블 요소를 헤더 + 데이터 행으로 작성한다."""
    columns = content.columns
    num_cols = len(columns)

    # 캡션이 있으면 먼저 작성
    if content.caption:
        cell = ws.cell(row=current_row, column=1, value=content.caption)
        cell.font = Font(bold=True, size=10, color="333333")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if num_cols > 1:
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=num_cols,
            )
        current_row += 1

    # 열 너비 설정
    col_widths = _calc_column_widths(columns)
    for col_idx, width in enumerate(col_widths, 1):
        col_letter = get_column_letter(col_idx)
        # 기존 너비보다 넓으면 갱신
        existing = ws.column_dimensions[col_letter].width or 0
        if width > existing:
            ws.column_dimensions[col_letter].width = width

    # 헤더 행
    header_row = current_row
    for col_idx, col_def in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_def.name)
        cell.font = _TABLE_HEADER_FONT
        cell.fill = _TABLE_HEADER_FILL
        cell.alignment = _TABLE_HEADER_ALIGN
        cell.border = _THIN_BORDER
    current_row += 1

    # 데이터 행
    for row_idx, row_data in enumerate(content.rows):
        for col_idx, col_def in enumerate(columns):
            value = row_data[col_idx] if col_idx < len(row_data) else ""
            cell = ws.cell(row=current_row, column=col_idx + 1, value=value)
            cell.alignment = _ALIGN_MAP.get(col_def.align, _DATA_ALIGN_LEFT)
            cell.border = _THIN_BORDER
            # 교차 행 배경색 (짝수 행: 흰색, 홀수 행: 연한 회색)
            cell.fill = _DATA_FILL_ODD if row_idx % 2 == 1 else _DATA_FILL_EVEN
        current_row += 1

    # 셀 병합 처리
    if content.merge:
        for m in content.merge:
            # merge 좌표는 테이블 내 상대 좌표 → 절대 좌표로 변환
            abs_start_row = header_row + m.start_row
            abs_end_row = header_row + m.end_row
            abs_start_col = m.start_col
            abs_end_col = m.end_col
            ws.merge_cells(
                start_row=abs_start_row, start_column=abs_start_col,
                end_row=abs_end_row, end_column=abs_end_col,
            )
            # 병합된 셀에 값 설정
            if m.value:
                ws.cell(row=abs_start_row, column=abs_start_col, value=m.value)

    return current_row


def _write_list(ws, content: ListContent, current_row: int, max_cols: int) -> int:
    """목록 요소를 각 항목을 행으로 작성한다."""
    for idx, item in enumerate(content.items):
        if content.list_type == "numbered":
            prefix = f"{idx + 1}. "
        else:
            prefix = "\u2022 "  # bullet: •
        cell = ws.cell(row=current_row, column=1, value=f"{prefix}{item}")
        cell.alignment = _LIST_ALIGN
        cell.font = Font(size=10)
        if max_cols > 1:
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=max_cols,
            )
        current_row += 1
    return current_row


def _get_max_columns(doc: DocumentOutput) -> int:
    """문서 전체에서 최대 열 수를 계산한다 (제목 병합 범위 결정용)."""
    max_cols = 1
    for section in doc.sections:
        for elem in section.elements:
            if elem.type == "table" and isinstance(elem.content, TableContent):
                max_cols = max(max_cols, len(elem.content.columns))
    return max_cols


async def build_xlsx(doc: DocumentOutput, output_path: str = None) -> dict:
    """DocumentOutput JSON → XLSX 파일 생성.

    각 섹션의 table 요소를 Excel 시트에 배치.
    heading/paragraph는 시트 상단에 텍스트로 배치.

    Args:
        doc: DocumentOutput 스키마 객체
        output_path: 출력 파일 경로. None이면 tempfile로 생성.

    Returns:
        {"status": "success", "file_path": output_path, "message": "..."}
        또는 {"status": "error", "message": "..."}
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = (doc.title or "Sheet1")[:31]

        # 문서 전체 최대 열 수 (제목/heading 병합 범위)
        max_cols = _get_max_columns(doc)

        # 문서 제목
        current_row = _write_title(ws, doc.title, max_cols, current_row=1)
        # 제목 아래 빈 행
        current_row += 1

        # 각 섹션 순차 배치
        for sec_idx, section in enumerate(doc.sections):
            # 섹션 제목 (heading으로 처리)
            if section.section_title:
                section_heading = TextContent(text=section.section_title, bold=True)
                current_row = _write_heading(ws, section_heading, current_row, max_cols)

            # 섹션 내 요소들
            for elem in section.elements:
                if elem.type == "heading" and isinstance(elem.content, TextContent):
                    current_row = _write_heading(ws, elem.content, current_row, max_cols)
                elif elem.type == "paragraph" and isinstance(elem.content, TextContent):
                    current_row = _write_paragraph(ws, elem.content, current_row, max_cols)
                elif elem.type == "table" and isinstance(elem.content, TableContent):
                    current_row = _write_table(ws, elem.content, current_row)
                elif elem.type == "list" and isinstance(elem.content, ListContent):
                    current_row = _write_list(ws, elem.content, current_row, max_cols)

            # 섹션 사이 빈 행 1개
            current_row += 1

        # 파일 저장
        if not output_path:
            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", prefix="doc_", delete=False,
            )
            output_path = tmp.name
            tmp.close()

        wb.save(output_path)
        wb.close()

        logger.info(f"[XLSX Builder] 생성 완료: {output_path}")
        return {
            "status": "success",
            "file_path": output_path,
            "message": f"XLSX 파일 생성 완료: {output_path}",
        }

    except Exception as e:
        error_msg = f"XLSX 생성 중 오류: {e}\n{traceback.format_exc()}"
        logger.error(f"[XLSX Builder] {error_msg}")
        return {"status": "error", "message": error_msg}
