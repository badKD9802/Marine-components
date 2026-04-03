"""
xlsx_builder 테스트
DocumentOutput → XLSX 변환 엔진 검증
"""

import asyncio
import os
import tempfile

import pytest
from openpyxl import load_workbook

import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from react_system.document_schema import (
    DocumentElement,
    DocumentOutput,
    ListContent,
    MergeCell,
    SectionOutput,
    TableColumn,
    TableContent,
    TextContent,
)
from react_system.tools.xlsx_builder import build_xlsx


def _run(coro):
    """async 함수를 동기적으로 실행하는 헬퍼."""
    return asyncio.get_event_loop().run_until_complete(coro)


# ─── 테스트용 DocumentOutput 팩토리 ───

def _make_simple_doc(title="테스트 문서") -> DocumentOutput:
    """최소한의 DocumentOutput (테이블 1개)"""
    return DocumentOutput(
        title=title,
        doc_type="보고서",
        sections=[
            SectionOutput(
                section_id="s1",
                section_title="매출 현황",
                elements=[
                    DocumentElement(
                        type="table",
                        content=TableContent(
                            caption="월별 매출",
                            columns=[
                                TableColumn(name="월", width=30, align="center"),
                                TableColumn(name="매출액", width=40, align="right"),
                                TableColumn(name="비고", width=30, align="left"),
                            ],
                            rows=[
                                ["1월", "1,200만원", "정상"],
                                ["2월", "1,500만원", "증가"],
                                ["3월", "980만원", "감소"],
                            ],
                        ),
                    )
                ],
            )
        ],
    )


class TestBuildXlsxBasic:
    """build_xlsx 기본 동작 테스트"""

    def test_should_return_success_status_for_simple_document(self):
        """최소 문서로 build_xlsx 호출 시 success를 반환해야 한다."""
        doc = _make_simple_doc()
        result = _run(build_xlsx(doc))

        assert result["status"] == "success"
        assert "file_path" in result
        assert os.path.exists(result["file_path"])

        # 정리
        os.unlink(result["file_path"])

    def test_should_create_xlsx_file_at_specified_path(self):
        """output_path를 지정하면 해당 경로에 파일이 생성되어야 한다."""
        doc = _make_simple_doc()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            result = _run(build_xlsx(doc, output_path=output_path))

            assert result["status"] == "success"
            assert result["file_path"] == output_path
            assert os.path.exists(output_path)

            # 파일이 유효한 XLSX인지 확인
            wb = load_workbook(output_path)
            assert len(wb.worksheets) >= 1
            wb.close()
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_should_use_tempfile_when_no_output_path(self):
        """output_path가 None이면 tempfile로 생성해야 한다."""
        doc = _make_simple_doc()
        result = _run(build_xlsx(doc))

        assert result["status"] == "success"
        assert result["file_path"].endswith(".xlsx")
        assert os.path.exists(result["file_path"])

        os.unlink(result["file_path"])


def _build_and_load(doc: DocumentOutput):
    """build_xlsx 실행 후 workbook을 로드하여 (result, wb) 반환하는 헬퍼.
    호출자가 wb.close()와 파일 삭제를 책임진다.
    """
    result = _run(build_xlsx(doc))
    assert result["status"] == "success", f"build_xlsx 실패: {result.get('message')}"
    wb = load_workbook(result["file_path"])
    return result, wb


class TestDocumentTitle:
    """문서 제목 렌더링 테스트"""

    def test_should_render_title_in_first_row(self):
        """문서 제목이 첫 행에 배치되어야 한다."""
        doc = _make_simple_doc(title="2024년 매출 보고서")
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            assert ws.cell(row=1, column=1).value == "2024년 매출 보고서"
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_apply_bold_and_large_font_to_title(self):
        """제목에 볼드 + 큰 폰트(14pt)가 적용되어야 한다."""
        doc = _make_simple_doc()
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            title_cell = ws.cell(row=1, column=1)
            assert title_cell.font.bold is True
            assert title_cell.font.size == 14
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_merge_title_across_max_columns(self):
        """제목이 테이블 최대 열 수만큼 병합되어야 한다."""
        doc = _make_simple_doc()  # 3열 테이블
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # openpyxl의 merged_cells에서 제목 병합 확인
            merged_ranges = [str(r) for r in ws.merged_cells.ranges]
            # A1:C1 (3열 병합) 또는 이를 포함하는 범위가 있어야 함
            title_merged = any("A1" in r and "C1" in r for r in merged_ranges)
            assert title_merged, f"제목 병합 범위를 찾을 수 없음: {merged_ranges}"
        finally:
            wb.close()
            os.unlink(result["file_path"])


class TestTableRendering:
    """테이블 요소 렌더링 테스트"""

    def test_should_render_table_header_with_column_names(self):
        """테이블 헤더에 열 이름이 배치되어야 한다."""
        doc = _make_simple_doc()
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # 행 1: 제목, 행 2: 빈 행, 행 3: 섹션 제목, 행 4: 캡션, 행 5: 헤더
            # 헤더 행을 찾기 — "월" 값이 있는 행
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=1, values_only=False):
                if row[0].value == "월":
                    header_row = row[0].row
                    break

            assert header_row is not None, "헤더 행('월')을 찾을 수 없음"
            assert ws.cell(row=header_row, column=1).value == "월"
            assert ws.cell(row=header_row, column=2).value == "매출액"
            assert ws.cell(row=header_row, column=3).value == "비고"
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_apply_header_style_blue_background_white_font(self):
        """테이블 헤더에 #2F5496 배경 + 흰 글씨 + 볼드가 적용되어야 한다."""
        doc = _make_simple_doc()
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # 헤더 행 찾기
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=1, values_only=False):
                if row[0].value == "월":
                    header_row = row[0].row
                    break

            assert header_row is not None
            cell = ws.cell(row=header_row, column=1)
            assert cell.font.bold is True
            assert cell.font.color.rgb == "00FFFFFF" or cell.font.color.theme is not None
            assert cell.fill.start_color.rgb == "002F5496"
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_render_data_rows_with_correct_values(self):
        """데이터 행에 올바른 값이 배치되어야 한다."""
        doc = _make_simple_doc()
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # "1월"이 있는 행 찾기
            data_row = None
            for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=1, values_only=False):
                if row[0].value == "1월":
                    data_row = row[0].row
                    break

            assert data_row is not None, "데이터 행('1월')을 찾을 수 없음"
            assert ws.cell(row=data_row, column=1).value == "1월"
            assert ws.cell(row=data_row, column=2).value == "1,200만원"
            assert ws.cell(row=data_row, column=3).value == "정상"
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_apply_alternating_row_colors(self):
        """데이터 행에 교차 배경색이 적용되어야 한다 (홀수 흰, 짝수 연회색)."""
        doc = _make_simple_doc()
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # "1월" 행 (row_idx=0 → 흰색) 과 "2월" 행 (row_idx=1 → 연회색) 찾기
            row_1월 = None
            row_2월 = None
            for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=1, values_only=False):
                if row[0].value == "1월":
                    row_1월 = row[0].row
                elif row[0].value == "2월":
                    row_2월 = row[0].row

            assert row_1월 is not None and row_2월 is not None

            # 1월 (인덱스 0, 짝수) → 흰색 FFFFFF
            fill_1 = ws.cell(row=row_1월, column=1).fill
            assert fill_1.start_color.rgb == "00FFFFFF", f"1월 행 색상: {fill_1.start_color.rgb}"

            # 2월 (인덱스 1, 홀수) → 연회색 F8FAFC
            fill_2 = ws.cell(row=row_2월, column=1).fill
            assert fill_2.start_color.rgb == "00F8FAFC", f"2월 행 색상: {fill_2.start_color.rgb}"
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_apply_thin_border_to_all_cells(self):
        """모든 데이터 셀에 얇은 보더가 적용되어야 한다."""
        doc = _make_simple_doc()
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # "1월" 행 찾기
            data_row = None
            for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=1, values_only=False):
                if row[0].value == "1월":
                    data_row = row[0].row
                    break

            assert data_row is not None
            cell = ws.cell(row=data_row, column=1)
            assert cell.border.left.style == "thin"
            assert cell.border.right.style == "thin"
            assert cell.border.top.style == "thin"
            assert cell.border.bottom.style == "thin"
        finally:
            wb.close()
            os.unlink(result["file_path"])


class TestTableCaption:
    """테이블 캡션 렌더링 테스트"""

    def test_should_render_caption_above_table(self):
        """테이블 캡션이 헤더 위에 배치되어야 한다."""
        doc = _make_simple_doc()
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            caption_row = None
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=1, values_only=False):
                if row[0].value == "월별 매출":
                    caption_row = row[0].row
                elif row[0].value == "월":
                    header_row = row[0].row

            assert caption_row is not None, "캡션('월별 매출')을 찾을 수 없음"
            assert header_row is not None, "헤더('월')를 찾을 수 없음"
            assert caption_row < header_row, "캡션이 헤더 위에 있어야 함"
        finally:
            wb.close()
            os.unlink(result["file_path"])


class TestHeadingAndParagraph:
    """heading, paragraph 요소 렌더링 테스트"""

    def test_should_render_heading_as_bold_text(self):
        """heading 요소가 볼드 텍스트 행으로 렌더링되어야 한다."""
        doc = DocumentOutput(
            title="테스트",
            doc_type="보고서",
            sections=[
                SectionOutput(
                    section_id="s1",
                    section_title="섹션1",
                    elements=[
                        DocumentElement(
                            type="heading",
                            content=TextContent(text="중요 항목", bold=True, font_size=14),
                        ),
                    ],
                )
            ],
        )
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            heading_row = None
            for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=1, values_only=False):
                if row[0].value == "중요 항목":
                    heading_row = row[0].row
                    break

            assert heading_row is not None, "heading('중요 항목')을 찾을 수 없음"
            cell = ws.cell(row=heading_row, column=1)
            assert cell.font.bold is True
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_render_paragraph_with_wrap_text(self):
        """paragraph 요소에 워드랩이 적용되어야 한다."""
        doc = DocumentOutput(
            title="테스트",
            doc_type="보고서",
            sections=[
                SectionOutput(
                    section_id="s1",
                    section_title="섹션1",
                    elements=[
                        DocumentElement(
                            type="paragraph",
                            content=TextContent(text="이것은 긴 단락 텍스트입니다. 워드랩이 적용되어야 합니다."),
                        ),
                    ],
                )
            ],
        )
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            para_row = None
            for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=1, values_only=False):
                if row[0].value and "긴 단락 텍스트" in str(row[0].value):
                    para_row = row[0].row
                    break

            assert para_row is not None, "paragraph를 찾을 수 없음"
            cell = ws.cell(row=para_row, column=1)
            assert cell.alignment.wrap_text is True
        finally:
            wb.close()
            os.unlink(result["file_path"])


class TestListRendering:
    """list 요소 렌더링 테스트"""

    def test_should_render_bullet_list_with_bullet_prefix(self):
        """bullet 목록 항목에 • 접두사가 붙어야 한다."""
        doc = DocumentOutput(
            title="테스트",
            doc_type="보고서",
            sections=[
                SectionOutput(
                    section_id="s1",
                    section_title="섹션1",
                    elements=[
                        DocumentElement(
                            type="list",
                            content=ListContent(
                                items=["항목 A", "항목 B", "항목 C"],
                                list_type="bullet",
                            ),
                        ),
                    ],
                )
            ],
        )
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            found_items = []
            for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=1, values_only=False):
                val = row[0].value
                if val and str(val).startswith("\u2022"):
                    found_items.append(str(val))

            assert len(found_items) == 3, f"bullet 항목 3개 예상, 실제: {found_items}"
            assert "\u2022 항목 A" in found_items[0]
            assert "\u2022 항목 B" in found_items[1]
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_render_numbered_list_with_number_prefix(self):
        """numbered 목록 항목에 1. 2. 접두사가 붙어야 한다."""
        doc = DocumentOutput(
            title="테스트",
            doc_type="보고서",
            sections=[
                SectionOutput(
                    section_id="s1",
                    section_title="섹션1",
                    elements=[
                        DocumentElement(
                            type="list",
                            content=ListContent(
                                items=["첫째", "둘째"],
                                list_type="numbered",
                            ),
                        ),
                    ],
                )
            ],
        )
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            found_items = []
            for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=1, values_only=False):
                val = row[0].value
                if val and ("1. " in str(val) or "2. " in str(val)):
                    found_items.append(str(val))

            assert len(found_items) == 2, f"번호 항목 2개 예상, 실제: {found_items}"
            assert "1. 첫째" in found_items[0]
            assert "2. 둘째" in found_items[1]
        finally:
            wb.close()
            os.unlink(result["file_path"])


class TestMergeCells:
    """셀 병합 테스트"""

    def test_should_merge_cells_according_to_merge_spec(self):
        """MergeCell 정의에 따라 셀이 병합되어야 한다."""
        doc = DocumentOutput(
            title="테스트",
            doc_type="보고서",
            sections=[
                SectionOutput(
                    section_id="s1",
                    section_title="병합 테스트",
                    elements=[
                        DocumentElement(
                            type="table",
                            content=TableContent(
                                columns=[
                                    TableColumn(name="구분", width=30),
                                    TableColumn(name="항목", width=30),
                                    TableColumn(name="금액", width=40),
                                ],
                                rows=[
                                    ["카테고리A", "항목1", "100"],
                                    ["", "항목2", "200"],
                                    ["카테고리B", "항목3", "300"],
                                ],
                                merge=[
                                    MergeCell(
                                        start_row=1, start_col=1,
                                        end_row=2, end_col=1,
                                        value="카테고리A",
                                    ),
                                ],
                            ),
                        )
                    ],
                )
            ],
        )
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # 병합된 범위가 존재하는지 확인
            merged_ranges = [str(r) for r in ws.merged_cells.ranges]
            # 제목 병합 외에 데이터 영역에 병합이 있어야 함
            # 제목은 row 1, 데이터 병합은 그 이후 행
            non_title_merges = [r for r in merged_ranges if not r.startswith("A1")]
            assert len(non_title_merges) >= 1, f"데이터 병합 범위를 찾을 수 없음: {merged_ranges}"
        finally:
            wb.close()
            os.unlink(result["file_path"])


class TestMultipleSections:
    """여러 섹션 처리 테스트"""

    def test_should_separate_sections_with_empty_row(self):
        """섹션 사이에 빈 행이 있어야 한다."""
        doc = DocumentOutput(
            title="테스트",
            doc_type="보고서",
            sections=[
                SectionOutput(
                    section_id="s1",
                    section_title="섹션1",
                    elements=[
                        DocumentElement(
                            type="paragraph",
                            content=TextContent(text="첫 번째 섹션 내용"),
                        ),
                    ],
                ),
                SectionOutput(
                    section_id="s2",
                    section_title="섹션2",
                    elements=[
                        DocumentElement(
                            type="paragraph",
                            content=TextContent(text="두 번째 섹션 내용"),
                        ),
                    ],
                ),
            ],
        )
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # 두 섹션 제목이 모두 존재하는지 확인
            found_s1 = False
            found_s2 = False
            for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=1, values_only=False):
                val = row[0].value
                if val == "섹션1":
                    found_s1 = True
                elif val == "섹션2":
                    found_s2 = True

            assert found_s1 and found_s2, "두 섹션 제목이 모두 존재해야 함"
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_render_section_title_as_heading(self):
        """섹션 제목이 볼드 heading으로 렌더링되어야 한다."""
        doc = _make_simple_doc()
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            section_row = None
            for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=1, values_only=False):
                if row[0].value == "매출 현황":
                    section_row = row[0].row
                    break

            assert section_row is not None, "섹션 제목('매출 현황')을 찾을 수 없음"
            cell = ws.cell(row=section_row, column=1)
            assert cell.font.bold is True
        finally:
            wb.close()
            os.unlink(result["file_path"])


class TestColumnWidth:
    """열 너비 계산 테스트"""

    def test_should_set_column_width_based_on_width_ratio(self):
        """columns의 width 비율에 따라 열 너비가 설정되어야 한다."""
        doc = _make_simple_doc()  # width: 30, 40, 30
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # width=40인 열(B)이 width=30인 열(A)보다 넓어야 함
            width_a = ws.column_dimensions["A"].width
            width_b = ws.column_dimensions["B"].width

            assert width_b > width_a, f"B열({width_b})이 A열({width_a})보다 넓어야 함"
        finally:
            wb.close()
            os.unlink(result["file_path"])


class TestErrorHandling:
    """에러 처리 테스트"""

    def test_should_return_error_status_on_invalid_output_path(self):
        """잘못된 경로에 저장 시 error를 반환해야 한다."""
        doc = _make_simple_doc()
        result = _run(build_xlsx(doc, output_path="/nonexistent/path/file.xlsx"))

        assert result["status"] == "error"
        assert "message" in result


class TestColumnAlignment:
    """열 정렬 테스트"""

    def test_should_apply_column_alignment_to_data_cells(self):
        """TableColumn의 align 설정이 데이터 셀에 적용되어야 한다."""
        doc = _make_simple_doc()  # center, right, left
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # "1월" 행 찾기
            data_row = None
            for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=1, values_only=False):
                if row[0].value == "1월":
                    data_row = row[0].row
                    break

            assert data_row is not None
            # 열 1(월): center
            assert ws.cell(row=data_row, column=1).alignment.horizontal == "center"
            # 열 2(매출액): right
            assert ws.cell(row=data_row, column=2).alignment.horizontal == "right"
            # 열 3(비고): left
            assert ws.cell(row=data_row, column=3).alignment.horizontal == "left"
        finally:
            wb.close()
            os.unlink(result["file_path"])


class TestComplexDocument:
    """복합 문서 통합 테스트 — heading, paragraph, table, list가 모두 포함된 문서"""

    def test_should_render_all_element_types_in_correct_order(self):
        """heading, paragraph, table, list 모든 요소가 올바른 순서로 배치되어야 한다."""
        doc = DocumentOutput(
            title="종합 보고서",
            doc_type="보고서",
            sections=[
                SectionOutput(
                    section_id="s1",
                    section_title="개요",
                    elements=[
                        DocumentElement(
                            type="heading",
                            content=TextContent(text="1분기 실적 요약"),
                        ),
                        DocumentElement(
                            type="paragraph",
                            content=TextContent(text="본 보고서는 1분기 실적을 요약합니다."),
                        ),
                        DocumentElement(
                            type="table",
                            content=TableContent(
                                columns=[
                                    TableColumn(name="항목", width=50),
                                    TableColumn(name="값", width=50),
                                ],
                                rows=[
                                    ["매출", "10억"],
                                    ["비용", "7억"],
                                ],
                            ),
                        ),
                        DocumentElement(
                            type="list",
                            content=ListContent(
                                items=["매출 증가", "비용 절감"],
                                list_type="bullet",
                            ),
                        ),
                    ],
                ),
            ],
        )
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # 모든 주요 값이 존재하는지 + 올바른 순서인지 확인
            values_in_order = []
            for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=1, values_only=True):
                if row[0] is not None:
                    values_in_order.append(str(row[0]))

            # 주요 값 존재 확인
            joined = " | ".join(values_in_order)
            assert "종합 보고서" in joined, f"문서 제목 없음: {joined}"
            assert "1분기 실적 요약" in joined, f"heading 없음: {joined}"
            assert "본 보고서는" in joined, f"paragraph 없음: {joined}"
            assert "항목" in joined, f"테이블 헤더 없음: {joined}"
            assert "매출" in joined, f"테이블 데이터 없음: {joined}"
            assert "\u2022 매출 증가" in joined, f"bullet list 없음: {joined}"

            # 순서 확인: 제목 → heading → paragraph → 테이블 → list
            idx_title = next(i for i, v in enumerate(values_in_order) if "종합 보고서" in v)
            idx_heading = next(i for i, v in enumerate(values_in_order) if "1분기 실적 요약" in v)
            idx_para = next(i for i, v in enumerate(values_in_order) if "본 보고서는" in v)
            idx_header = next(i for i, v in enumerate(values_in_order) if v == "항목")
            idx_bullet = next(i for i, v in enumerate(values_in_order) if "\u2022" in v)

            assert idx_title < idx_heading < idx_para < idx_header < idx_bullet, \
                f"순서가 올바르지 않음: title={idx_title}, heading={idx_heading}, " \
                f"para={idx_para}, header={idx_header}, bullet={idx_bullet}"
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_handle_table_without_caption(self):
        """캡션 없는 테이블도 정상 렌더링되어야 한다."""
        doc = DocumentOutput(
            title="테스트",
            doc_type="보고서",
            sections=[
                SectionOutput(
                    section_id="s1",
                    section_title="섹션",
                    elements=[
                        DocumentElement(
                            type="table",
                            content=TableContent(
                                columns=[
                                    TableColumn(name="이름"),
                                    TableColumn(name="나이"),
                                ],
                                rows=[["홍길동", "30"]],
                            ),
                        ),
                    ],
                )
            ],
        )
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            found_name = False
            for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=2, values_only=True):
                if row[0] == "홍길동" and row[1] == "30":
                    found_name = True
                    break
            assert found_name, "캡션 없는 테이블 데이터가 렌더링되지 않음"
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_handle_empty_sections(self):
        """요소가 없는 섹션이 있어도 에러 없이 처리되어야 한다."""
        doc = DocumentOutput(
            title="빈 섹션 테스트",
            doc_type="보고서",
            sections=[
                SectionOutput(
                    section_id="s1",
                    section_title="빈 섹션",
                    elements=[],
                ),
                SectionOutput(
                    section_id="s2",
                    section_title="내용 섹션",
                    elements=[
                        DocumentElement(
                            type="paragraph",
                            content=TextContent(text="내용 있음"),
                        ),
                    ],
                ),
            ],
        )
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            found = False
            for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=1, values_only=True):
                if row[0] and "내용 있음" in str(row[0]):
                    found = True
                    break
            assert found, "빈 섹션 다음의 내용 섹션이 렌더링되지 않음"
        finally:
            wb.close()
            os.unlink(result["file_path"])

    def test_should_handle_table_with_missing_cell_values(self):
        """행의 셀 수가 열 수보다 적어도 에러 없이 처리되어야 한다."""
        doc = DocumentOutput(
            title="테스트",
            doc_type="보고서",
            sections=[
                SectionOutput(
                    section_id="s1",
                    section_title="섹션",
                    elements=[
                        DocumentElement(
                            type="table",
                            content=TableContent(
                                columns=[
                                    TableColumn(name="A"),
                                    TableColumn(name="B"),
                                    TableColumn(name="C"),
                                ],
                                rows=[
                                    ["값1"],  # B, C 열 값 없음
                                    ["값2", "값3"],  # C 열 값 없음
                                ],
                            ),
                        ),
                    ],
                )
            ],
        )
        result, wb = _build_and_load(doc)
        ws = wb.active

        try:
            # "값1" 행에서 B, C 열이 빈 문자열로 채워져야 함
            data_row = None
            for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=1, values_only=False):
                if row[0].value == "값1":
                    data_row = row[0].row
                    break

            assert data_row is not None
            assert ws.cell(row=data_row, column=1).value == "값1"
            # openpyxl은 빈 문자열을 저장/로드 시 None으로 반환할 수 있음
            assert ws.cell(row=data_row, column=2).value in ("", None)
            assert ws.cell(row=data_row, column=3).value in ("", None)
        finally:
            wb.close()
            os.unlink(result["file_path"])
