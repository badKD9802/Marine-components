"""
Excel 도구
구조화된 데이터를 <excel-data> JSON 태그로 출력하여 ExcelReadonlyView로 렌더링합니다.

프론트엔드의 parseExcelData → ExcelStreamTable → ExcelReadonlyView 경로를 활용하며,
XLSX 다운로드는 클라이언트 사이드에서 처리합니다.
"""

import json
import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .html_format_tools import _AUTO_LABELS, _fmt_cell

# ── 합계/소계 행 키워드 ──
_SUMMARY_KEYWORDS = {"합계", "소계", "총계", "합산", "Total", "Subtotal", "TOTAL", "SUBTOTAL", "total", "subtotal"}

# ── 전화번호 패턴 (숫자 오탐 방지) ──
_PHONE_RE = re.compile(r"^\d{2,4}[-.\s]\d{3,4}[-.\s]\d{4}$")

# ── 조건부 서식: 상태 키워드 → 색상 ──
_STATUS_COLORS = {
    "완료": "006100", "성공": "006100", "승인": "006100", "정상": "006100",
    "진행중": "9C6500", "진행": "9C6500", "대기": "9C6500", "검토중": "9C6500", "보류": "9C6500",
    "미완료": "9C0006", "실패": "9C0006", "반려": "9C0006", "취소": "9C0006", "오류": "9C0006",
}

# ── 교차 행 배경색 ──
_ALT_FILL_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_ALT_FILL_ODD = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

# ── 합계 행 스타일 ──
_SUMMARY_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
_SUMMARY_FONT = Font(bold=True, size=11)
_SUMMARY_TOP_BORDER = Border(
    top=Side(style="medium", color="94A3B8"),
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def _detect_cell_type(value):
    """셀 값을 분석하여 (typed_value, number_format, is_numeric) 반환."""
    if value is None or value == "":
        return value, None, False

    # 이미 숫자인 경우
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != int(value):
            return value, "#,##0.00", True
        return value, "#,##0", True

    s = str(value).strip()
    if not s:
        return value, None, False

    # 전화번호 → 문자열 유지
    if _PHONE_RE.match(s):
        return s, None, False

    # 선행 0 (사번, 코드) → 문자열 유지
    if len(s) >= 2 and s[0] == "0" and s[1].isdigit() and "." not in s:
        return s, None, False

    # 퍼센트 ("12.5%")
    if s.endswith("%"):
        try:
            num = float(s[:-1].replace(",", ""))
            return num / 100, "0.0%", True
        except ValueError:
            pass

    # 통화/콤마 숫자 ("1,234,000", "1,234원", "$1,234")
    cleaned = re.sub(r"[원₩$,\s]", "", s)
    if cleaned and cleaned not in s:  # 실제로 치환이 발생한 경우
        try:
            num = float(cleaned)
            if num == int(num):
                return int(num), "#,##0", True
            return num, "#,##0.00", True
        except ValueError:
            pass

    # 날짜 ("2025-03-08", "2025.03.08")
    date_match = re.match(r"^(\d{4})[-./ ](\d{1,2})[-./ ](\d{1,2})$", s)
    if date_match:
        try:
            dt = datetime(int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3)))
            return dt, "yyyy-mm-dd", False
        except ValueError:
            pass

    # 순수 숫자 ("1234", "-56.78")
    try:
        num = float(s)
        if "." in s:
            return num, "#,##0.00", True
        return int(num), "#,##0", True
    except ValueError:
        pass

    return value, None, False


def _is_summary_row(row, columns):
    """첫 번째 열 값이 합계/소계 키워드인지 확인."""
    if not columns:
        return False
    first_val = str(row.get(columns[0], "")).strip()
    return first_val in _SUMMARY_KEYWORDS


def format_data_as_excel(title, data, columns=None, column_labels=None, file_name=None, _auth=None, header_groups=None, format_rules=None):
    """구조화된 데이터를 <excel-data> JSON 태그로 출력합니다.

    프론트엔드의 parseExcelData → ExcelStreamTable → ExcelReadonlyView 경로로
    스프레드시트 UI 렌더링 + 클라이언트 사이드 XLSX 다운로드를 제공합니다.

    Args:
        title: 테이블/파일 제목 (예: "월별 매출 현황")
        data: dict 리스트 (예: [{"월": "1월", "매출": "12억"}, ...])
        columns: 표시할 컬럼 키 목록 (순서대로). 생략 시 data[0]의 키를 자동 사용.
        column_labels: 컬럼 키 → 한글 라벨 매핑. 생략 시 자동 매핑.
        file_name: 다운로드 파일명 (확장자 제외). 생략 시 제목_날짜 자동 생성.

    Returns:
        dict: {"status": "success", "html_content": str, "text_summary": str}
    """
    try:
        # 1. 입력 검증
        if not data or not isinstance(data, list):
            return {"status": "error", "message": "data가 비어있거나 리스트가 아닙니다."}

        rows = [row for row in data if isinstance(row, dict)]
        if not rows:
            return {"status": "error", "message": "data에 유효한 dict 항목이 없습니다."}

        # 2. 컬럼/라벨 결정
        if not columns:
            seen = set()
            columns = []
            for row in rows:
                for k in row.keys():
                    if k not in seen:
                        seen.add(k)
                        columns.append(k)

        labels = dict(column_labels) if column_labels else {}
        for col in columns:
            if col not in labels:
                labels[col] = _AUTO_LABELS.get(col, col)

        # 3. 데이터 키를 라벨로 변환 (emp_nm → 이름)
        labeled_rows = []
        for row in rows:
            labeled_row = {}
            for col in columns:
                labeled_row[labels.get(col, col)] = row.get(col, "")
            labeled_rows.append(labeled_row)

        # 4. header_groups, format_rules의 컬럼 참조도 라벨로 변환
        labeled_hg = None
        if header_groups:
            labeled_hg = []
            for g in header_groups:
                labeled_hg.append({
                    "label": g.get("label", ""),
                    "columns": [labels.get(c, c) for c in g.get("columns", [])],
                })

        labeled_fr = None
        if format_rules:
            labeled_fr = []
            for r in format_rules:
                new_r = dict(r)
                new_r["column"] = labels.get(r.get("column", ""), r.get("column", ""))
                labeled_fr.append(new_r)

        # 5. 파일명 결정
        if not file_name:
            date_str = datetime.now().strftime("%Y%m%d")
            file_name = f"{title or '데이터'}_{date_str}"
        if file_name.endswith(".xlsx"):
            file_name = file_name[:-5]

        # 6. <excel-data> JSON 빌드
        excel_obj = {"title": title or "조회 결과", "data": labeled_rows, "file_name": file_name}
        if labeled_hg:
            excel_obj["header_groups"] = labeled_hg
        if labeled_fr:
            excel_obj["format_rules"] = labeled_fr

        html_content = f"<excel-data>\n{json.dumps(excel_obj, ensure_ascii=False)}\n</excel-data>"

        # 7. text_summary 생성
        text_summary = _build_text_summary(rows, columns, labels)

        return {
            "status": "success",
            "html_content": html_content,
            "text_summary": text_summary,
        }

    except Exception as e:
        import traceback
        error_msg = f"Excel 생성 중 오류:\n{str(e)}\n\n{traceback.format_exc()}"
        print(error_msg)
        return {"status": "error", "message": error_msg}


def _fill_worksheet(ws, rows, columns, labels, header_groups=None, format_rules=None):
    """워크시트에 헤더 + 데이터 행 채우기 (공통 헬퍼).

    Args:
        header_groups: [{"label": "그룹명", "columns": ["col1", "col2"]}, ...] (optional)
        format_rules: [{"column": "col", "type": "positive_negative|status|threshold", "value": N}, ...] (optional)
    """
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    group_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
    )

    # ── 다단 헤더 (header_groups) ──
    data_start_row = 2
    if header_groups:
        try:
            col_to_idx = {col: i for i, col in enumerate(columns)}
            # Row 1: 병합 그룹 헤더
            grouped_cols = set()
            for group in header_groups:
                g_cols = [c for c in group.get("columns", []) if c in col_to_idx]
                if not g_cols:
                    continue
                indices = sorted(col_to_idx[c] for c in g_cols)
                start_col = indices[0] + 1
                end_col = indices[-1] + 1
                if start_col != end_col:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws.cell(row=1, column=start_col, value=group.get("label", ""))
                cell.font = header_font
                cell.fill = group_fill
                cell.alignment = header_align
                cell.border = thin_border
                for ci in range(start_col, end_col + 1):
                    grouped_cols.add(ci)
                    if ci != start_col:
                        ws.cell(row=1, column=ci).border = thin_border

            # 그룹에 속하지 않는 열은 빈 셀 (스타일 적용)
            for col_idx in range(1, len(columns) + 1):
                if col_idx not in grouped_cols:
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = group_fill
                    cell.border = thin_border

            # Row 2: 개별 열 헤더
            for col_idx, col_key in enumerate(columns, 1):
                cell = ws.cell(row=2, column=col_idx, value=labels.get(col_key, col_key))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            data_start_row = 3
        except Exception:
            # 그룹 헤더 실패 → 단일 헤더 fallback
            header_groups = None

    if not header_groups:
        for col_idx, col_key in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=labels.get(col_key, col_key))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        data_start_row = 2

    # ── format_rules 인덱스 구성 ──
    rule_map = {}
    if format_rules:
        try:
            for rule in format_rules:
                col_name = rule.get("column", "")
                if col_name in columns:
                    rule_map[col_name] = rule
        except Exception:
            rule_map = {}

    # ── 합계 행 위치 사전 분석 (SUM 수식용) ──
    summary_indices = [i for i, row in enumerate(rows) if _is_summary_row(row, columns)]
    use_sum_formula = len(summary_indices) == 1 and summary_indices[0] == len(rows) - 1

    # ── 데이터 행 ──
    data_row_counter = 0  # 합계 행 제외한 데이터 행 번호
    for row_idx_offset, row in enumerate(rows):
        excel_row = data_start_row + row_idx_offset
        is_summary = _is_summary_row(row, columns)

        for col_idx, col_key in enumerate(columns, 1):
            val = row.get(col_key)
            typed_val, num_fmt, is_numeric = _detect_cell_type(val)

            # 합계 행 + SUM 수식 적용
            if is_summary and use_sum_formula and is_numeric and col_idx > 1:
                col_letter = get_column_letter(col_idx)
                formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{excel_row - 1})"
                cell = ws.cell(row=excel_row, column=col_idx, value=formula)
                cell.number_format = num_fmt or "#,##0"
            else:
                cell = ws.cell(row=excel_row, column=col_idx, value=typed_val if typed_val is not None else "")
                if num_fmt:
                    cell.number_format = num_fmt

            # 정렬: 숫자 → 우측, 문자열 → 좌측
            cell.alignment = Alignment(horizontal="right" if is_numeric else "left", vertical="center")

            # 합계 행 스타일
            if is_summary:
                cell.font = _SUMMARY_FONT
                cell.fill = _SUMMARY_FILL
                cell.border = _SUMMARY_TOP_BORDER
            else:
                cell.border = thin_border
                # 교차 행 배경색
                cell.fill = _ALT_FILL_ODD if data_row_counter % 2 == 1 else _ALT_FILL_EVEN

            # 조건부 서식 적용
            if col_key in rule_map and not is_summary:
                _apply_conditional_style(cell, typed_val, rule_map[col_key])

        if not is_summary:
            data_row_counter += 1

    # ── 틀 고정 (헤더 행 고정) ──
    ws.freeze_panes = ws.cell(row=data_start_row, column=1)

    # ── 자동 열 너비 ──
    for col_idx, col_key in enumerate(columns, 1):
        max_len = len(str(labels.get(col_key, col_key)))
        for row in rows[:50]:
            val_len = len(str(row.get(col_key, "")))
            if val_len > max_len:
                max_len = val_len
        # 그룹 헤더 길이도 고려
        if header_groups:
            for group in header_groups:
                g_cols = group.get("columns", [])
                if col_key in g_cols and len(g_cols) > 0:
                    group_len = len(group.get("label", "")) // len(g_cols) + 2
                    if group_len > max_len:
                        max_len = group_len
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)


def _apply_conditional_style(cell, typed_val, rule):
    """셀에 조건부 서식 적용."""
    rule_type = rule.get("type", "")
    try:
        if rule_type == "positive_negative":
            if isinstance(typed_val, (int, float)):
                if typed_val > 0:
                    cell.font = Font(color="006100", bold=cell.font.bold)
                elif typed_val < 0:
                    cell.font = Font(color="9C0006", bold=cell.font.bold)
        elif rule_type == "status":
            status_key = str(typed_val).strip()
            color = _STATUS_COLORS.get(status_key)
            if color:
                cell.font = Font(color=color, bold=True)
        elif rule_type == "threshold":
            threshold_val = rule.get("value", 0)
            if isinstance(typed_val, (int, float)):
                if typed_val >= threshold_val:
                    cell.font = Font(color="006100", bold=cell.font.bold)
                else:
                    cell.font = Font(color="9C0006", bold=cell.font.bold)
    except Exception:
        pass


def _fill_worksheet_sections(ws, sections):
    """여러 섹션(다른 컬럼 구조)을 하나의 시트에 세로로 배치.

    Args:
        sections: [{"subtitle": "구분명", "data": [dict, ...], "columns": [...](optional)}, ...]
    """
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    subtitle_font = Font(bold=True, color="FFFFFF", size=12)
    subtitle_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
    )

    current_row = 1
    col_widths = {}  # col_position → max_width

    for sec_idx, section in enumerate(sections):
        subtitle = section.get("subtitle", f"Section {sec_idx + 1}")
        data = section.get("data", [])
        rows = [r for r in data if isinstance(r, dict)]
        if not rows:
            continue

        # 컬럼 결정
        sec_columns = section.get("columns")
        if not sec_columns:
            seen = set()
            sec_columns = []
            for row in rows:
                for k in row.keys():
                    if k not in seen:
                        seen.add(k)
                        sec_columns.append(k)

        labels = {}
        for col in sec_columns:
            labels[col] = _AUTO_LABELS.get(col, col)
        num_cols = len(sec_columns)

        # ── 소제목 행 ──
        cell = ws.cell(row=current_row, column=1, value=subtitle)
        cell.font = subtitle_font
        cell.fill = subtitle_fill
        cell.alignment = Alignment(vertical="center")
        if num_cols > 1:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        for ci in range(1, num_cols + 1):
            ws.cell(row=current_row, column=ci).fill = subtitle_fill
        current_row += 1

        # ── 열 헤더 행 ──
        for col_idx, col_key in enumerate(sec_columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=labels.get(col_key, col_key))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        current_row += 1

        # ── 데이터 행 ──
        data_row_counter = 0
        for row in rows:
            is_summary = _is_summary_row(row, sec_columns)
            for col_idx, col_key in enumerate(sec_columns, 1):
                val = row.get(col_key)
                typed_val, num_fmt, is_numeric = _detect_cell_type(val)
                cell = ws.cell(row=current_row, column=col_idx, value=typed_val if typed_val is not None else "")
                if num_fmt:
                    cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right" if is_numeric else "left", vertical="center")
                if is_summary:
                    cell.font = _SUMMARY_FONT
                    cell.fill = _SUMMARY_FILL
                    cell.border = _SUMMARY_TOP_BORDER
                else:
                    cell.border = thin_border
                    cell.fill = _ALT_FILL_ODD if data_row_counter % 2 == 1 else _ALT_FILL_EVEN
            if not is_summary:
                data_row_counter += 1
            current_row += 1

        # 열 너비 추적
        for col_idx, col_key in enumerate(sec_columns, 1):
            w = len(str(labels.get(col_key, col_key)))
            for row in rows[:50]:
                vl = len(str(row.get(col_key, "")))
                if vl > w:
                    w = vl
            if col_idx not in col_widths or w > col_widths[col_idx]:
                col_widths[col_idx] = w

        # 섹션 간 빈 행
        current_row += 1

    # ── 자동 열 너비 ──
    for col_idx, max_len in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # ── 틀 고정 (첫 섹션 소제목 + 헤더 이후) ──
    if current_row > 1:
        ws.freeze_panes = ws.cell(row=3, column=1)


def _create_xlsx_workbook(rows, columns, labels, title, header_groups=None, format_rules=None):
    """단일 시트 워크북 생성."""
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Sheet1")[:31]
    _fill_worksheet(ws, rows, columns, labels, header_groups=header_groups, format_rules=format_rules)
    return wb


def _build_text_summary(rows, columns, labels):
    """LLM 참조용 텍스트 요약 생성 (최대 50행, 5열)."""
    text_lines = []
    display_cols = columns[:5]
    for row in rows[:50]:
        parts = [f"{labels.get(c, c)}: {_fmt_cell(c, row.get(c))}" for c in display_cols]
        text_lines.append("- " + " | ".join(parts))
    if len(rows) > 50:
        text_lines.append(f"... 외 {len(rows) - 50}건")
    return "\n".join(text_lines)


def save_xlsx_file(title, data, columns=None, column_labels=None, file_name=None, save_dir=None, header_groups=None, format_rules=None):
    """XLSX 파일을 디스크에 저장하고 파일명을 반환.

    Args:
        title: 테이블/파일 제목
        data: dict 리스트
        columns: 표시할 컬럼 키 목록. 생략 시 data[0]의 키를 자동 사용.
        column_labels: 컬럼 키 → 한글 라벨 매핑. 생략 시 자동 매핑.
        file_name: 다운로드 파일명 (확장자 제외). 생략 시 제목_날짜 자동 생성.
        save_dir: 저장 디렉토리 경로.

    Returns:
        dict: {"status": "success", "file_name": str, "file_path": str, "row_count": int}
    """
    try:
        if not data or not isinstance(data, list):
            return {"status": "error", "message": "data가 비어있거나 리스트가 아닙니다."}

        rows = [row for row in data if isinstance(row, dict)]
        if not rows:
            return {"status": "error", "message": "data에 유효한 dict 항목이 없습니다."}

        # 컬럼/라벨 결정
        if not columns:
            seen = set()
            columns = []
            for row in rows:
                for k in row.keys():
                    if k not in seen:
                        seen.add(k)
                        columns.append(k)

        labels = dict(column_labels) if column_labels else {}
        for col in columns:
            if col not in labels:
                labels[col] = _AUTO_LABELS.get(col, col)

        # 파일명 결정
        if not file_name:
            date_str = datetime.now().strftime("%Y%m%d")
            file_name = f"{title or '데이터'}_{date_str}"
        if not file_name.endswith(".xlsx"):
            file_name += ".xlsx"

        # 워크북 생성 및 저장
        wb = _create_xlsx_workbook(rows, columns, labels, title, header_groups=header_groups, format_rules=format_rules)
        os.makedirs(save_dir, exist_ok=True)
        file_path = os.path.join(save_dir, file_name)
        wb.save(file_path)
        wb.close()

        return {
            "status": "success",
            "file_name": file_name,
            "file_path": file_path,
            "row_count": len(rows),
        }

    except Exception as e:
        import traceback
        return {"status": "error", "message": f"Excel 저장 중 오류:\n{e}\n\n{traceback.format_exc()}"}


def save_multi_sheet_xlsx(sheets, file_name=None, save_dir=None):
    """여러 데이터를 시트별로 분리하여 하나의 XLSX 파일로 저장.

    Args:
        sheets: list of dicts, each with:
            - title: 시트 제목
            - data: dict 리스트
            - columns: (optional) 컬럼 키 목록
            - column_labels: (optional) 컬럼 키 → 라벨 매핑
        file_name: 파일명. 생략 시 자동 생성.
        save_dir: 저장 디렉토리.

    Returns:
        dict: {"status": "success", "file_name", "file_path", "row_count", "sheet_info"}
    """
    try:
        if not sheets:
            return {"status": "error", "message": "sheets가 비어있습니다."}

        wb = Workbook()
        wb.remove(wb.active)  # 기본 시트 제거
        total_rows = 0
        sheet_info = []

        for idx, sheet_data in enumerate(sheets):
            title = sheet_data.get("title", f"Sheet{idx + 1}")
            sections = sheet_data.get("sections")

            # 시트 이름 (31자 제한 + 중복 방지)
            sheet_title = (title or f"Sheet{idx + 1}")[:31]
            existing = [ws.title for ws in wb.worksheets]
            base_title = sheet_title
            counter = 2
            while sheet_title in existing:
                suffix = f" ({counter})"
                sheet_title = base_title[: 31 - len(suffix)] + suffix
                counter += 1

            if sections:
                # Sections 모드: 다른 컬럼 구조를 하나의 시트에 세로 배치
                valid_sections = [s for s in sections if any(isinstance(r, dict) for r in s.get("data", []))]
                if not valid_sections:
                    continue
                ws = wb.create_sheet(title=sheet_title)
                _fill_worksheet_sections(ws, valid_sections)
                sec_row_count = sum(len([r for r in s.get("data", []) if isinstance(r, dict)]) for s in valid_sections)
                total_rows += sec_row_count
                sheet_info.append({"title": title, "row_count": sec_row_count})
            else:
                # 기존 flat data 모드
                data = sheet_data.get("data", [])
                columns = sheet_data.get("columns")
                column_labels = sheet_data.get("column_labels")

                rows = [row for row in data if isinstance(row, dict)]
                if not rows:
                    continue

                if not columns:
                    seen = set()
                    columns = []
                    for row in rows:
                        for k in row.keys():
                            if k not in seen:
                                seen.add(k)
                                columns.append(k)

                labels = dict(column_labels) if column_labels else {}
                for col in columns:
                    if col not in labels:
                        labels[col] = _AUTO_LABELS.get(col, col)

                ws = wb.create_sheet(title=sheet_title)
                _fill_worksheet(
                    ws, rows, columns, labels,
                    header_groups=sheet_data.get("header_groups"),
                    format_rules=sheet_data.get("format_rules"),
                )
                total_rows += len(rows)
                sheet_info.append({"title": title, "row_count": len(rows)})

        if not wb.worksheets:
            return {"status": "error", "message": "유효한 데이터가 없습니다."}

        if not file_name:
            file_name = "data.xlsx"
        if not file_name.endswith(".xlsx"):
            file_name += ".xlsx"

        os.makedirs(save_dir, exist_ok=True)
        file_path = os.path.join(save_dir, file_name)
        wb.save(file_path)
        wb.close()

        return {
            "status": "success",
            "file_name": file_name,
            "file_path": file_path,
            "row_count": total_rows,
            "sheet_info": sheet_info,
        }

    except Exception as e:
        import traceback
        return {"status": "error", "message": f"Excel 저장 중 오류:\n{e}\n\n{traceback.format_exc()}"}
